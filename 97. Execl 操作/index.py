from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import math
# 获取文件
doc = load_workbook('医疗应收账款字段表.xlsx', data_only=True)

# 计算开始索引和结束索引
start_index = column_index_from_string('G') - 1
end_index = column_index_from_string('AT') - 1

# 客户名称索引
name_index = column_index_from_string('D') - 1

# 打开工作表
a_sheet = doc['历史新增应收账款表现']

huikuanTitle = [
    '1月内回款',
    '2月内回款',
    '3月内回款',
    '4月内回款',
    '5月内回款',
    '6月内回款',
    '7月内回款',
    '8月内回款',
    '9月内回款',
    '10月内回款',
    '11月内回款',
    '12月内回款',
    '13月内回款',
    '14月内回款',
    '15月内回款',
    '16月内回款',
    '17月内回款',
    '18月内回款',
    '19月内回款',
    '20月内回款',
    '21月内回款',
]

# 过滤数据表，以公司名字过滤
nameFilter = []


# 函数：获取日期列表 如 2017年5月
def getDateTitle ():
    sList = list(a_sheet.rows)[1]
    arr = {}
    for index in range(start_index, end_index, 2):
        arr[index] = sList[index].value
    return arr

# 函数：计算回款
def getHuikuan (add, reduce, name):
    data = []
    # 全部增加数
    allAdd = 0
    # 遍历增加，以每一个增加为一行数据
    for index, item in enumerate(add):
        # 查看公司是否在过滤名单中
        if name in nameFilter: break
        allAdd += item['add']
        # 全部减少数
        allReduce = 0
        # 状态
        status = False
        # 定义数据行
        row = {}
        # 设置公司名
        row['name'] = name
        # 当前增加的日期
        nowYear = item['date'].year
        nowMonth = item['date'].month
        # 设置当前月份
        row['date'] = '%s年%s月' % (nowYear, nowMonth)
        # 当期增加
        row['add'] = item['add']
        # 遍历减少，计算回款
        for rindex, reduceItem in enumerate(reduce): 
            allReduce += reduceItem['reduce']
            # 回款日期
            print(name, reduceItem)
            rYear = reduceItem['date'].year
            rMonth = reduceItem['date'].month
            # 计算回款间隔
            if rYear > nowYear:
                num = 12 - nowMonth + rMonth
            else:
                num = rMonth - nowMonth
                if num < 0 or rYear < nowYear: num = 0
            for n in range(0, num):
                if huikuanTitle[n] not in row:
                    row[huikuanTitle[n]] = 0
            # 回款差额
            nChae = allReduce - allAdd
            # 当减少数大于增加数的时候，回款率为1，此循 环结束
            if nChae >= 0:
                row[huikuanTitle[num]] = 1
                break
            elif nChae < 0 and nChae + item['add'] < 0:
                nHuikuanlv = 0
                row[huikuanTitle[num]] = nHuikuanlv
            else:
                # 保留6位小数
                nHuikuanlv = math.floor((1 + nChae/item['add']) * 1000000) / 1000000
                row[huikuanTitle[num]] = nHuikuanlv
        data.append(row)
    return data

# 获取表格的日期列表
dateTitle = getDateTitle()
# 新的数据表数据
newSheet = []

# 遍历源数据表
for row in list(a_sheet.rows)[3:312]:
    # 客户名称
    name = row[name_index].value
    # 存储本期增加的资金列表
    nowDateAdd = []
    # 存储本期减少的资金列表
    nowDateReduceNum = []
    # 遍历当前行中的资金数据
    for index in range(start_index, end_index, 2):
        # 取得本期增加
        addNum = row[index].value
        # 如果本期增加的数据是字符 - 或者没有数据，则设置为 0
        if addNum == '-' or not addNum:
            addNum = 0
        # 如果大于0则组合数据
        if addNum > 0:
            nowDateAdd.append({
                'date': dateTitle[index],
                'add': addNum
            })
        # 取得本期减少
        reduceNum = row[index + 1].value
        # 如果本期减少的数据是字符 - 或者没有数据，则设置为 0
        if reduceNum == '-' or not reduceNum:
            reduceNum = 0
        # 如果大于0则组合数据
        if reduceNum > 0:
            nowDateReduceNum.append({
                'date': dateTitle[index],
                'reduce': reduceNum
            })
    # 计算回款
    newSheet += getHuikuan(nowDateAdd, nowDateReduceNum, name)

# 创建新的表格
newXlsx = Workbook()
newXlsxS1 = newXlsx.active
newXlsxS1.title = '回款率统计表'

# 写表头
newXlsxS1['A1'] = '客户名称'
newXlsxS1['B1'] = '时间'
newXlsxS1['C1'] = '新增'

# 写表头，写入汇款日期表头
for index, item in enumerate(huikuanTitle):
    newXlsxS1.cell(row=1, column=index+4, value=item)

# 写数据，遍历回款数据数组写入数据
for index, row in enumerate(newSheet):
    for j, column in enumerate(row):
        # 从第二行开始写入
        newXlsxS1.cell(row=index+2, column=j+1, value=row[column])

# 保存文件
newXlsx.save(filename='回款率统计表.xlsx')


# 函数：获取之前的本期减少